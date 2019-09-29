#!/usr/bin/python3
# -*- coding: utf-8 -*-

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 폰트 설정
def setFont(cell):
    cell.font = Font(name='맑은 고딕', size=10, bold=True)

# 가운데 정렬
def setAlign(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 음영 색 지정
yellowFill = PatternFill(start_color='FFFFFF00',
            end_color='FFFFFF00',
            fill_type='solid')

def makeExcel(coinList):

    # 워크시트 생성
    wb = Workbook()
    ws = wb.active

    # 행 고정
    ws.freeze_panes = 'A4'  
    ws.merge_cells('A1:F1')
    ws['A1'] = 'ICO 현황'

    title = ['A3','B3','C3','D3','E3','F3']
    title_content = ['ID', 'NAME', 'ICO 시작일', 'ICO 종료일', 'ICO 총 상장액', '통화']

    # 제목 행 입력
    for i in range(len(title)):
        ws[title[i]] = title_content[i]

        # set the width of the column 
        ws.column_dimensions[title[i][:1]].width = 20
        
        # 지정된 음영 색으로 음역 색칠하기
        ws[title[i]].fill = yellowFill
        setFont(ws[title[i]])
        setAlign(ws[title[i]])

    for coin in coinList:
        # parsing
        print(coin)
        coin_id, coin_name, ico_sdate, ico_edate, ico_rasied, currency = coin.split('|')
        ws.append([coin_id, coin_name, ico_sdate[:10], ico_edate[:10], ico_rasied, currency])
    
    wb.save('test.xlsx')